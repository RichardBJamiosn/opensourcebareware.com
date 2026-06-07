"""
Generate downloadable XLSX templates for Open Source Barware.
Run: pip3 install openpyxl && python3 tools/generate_templates.py
Output: public/downloads/*.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "public", "downloads")
os.makedirs(OUT_DIR, exist_ok=True)

# ── Color palette (matches app theme) ────────────────────────────────────────
C_BG = "FF0C1018"
C_HEADER = "FF16202E"
C_COPPER = "FFA0B4CC"
C_TEXT = "FFE0E6EF"
C_MUTED = "FF8B95A8"
C_RED = "FF7C3A4E"
C_GREEN = "FF4A7C6F"
C_BORDER = "FF2A3A4A"
C_WHITE = "FFFFFFFF"
C_LIGHT_RED = "FFFDE8E8"
C_LIGHT_GREEN = "FFE8F5F0"

def _header_font():
    return Font(name="Calibri", bold=True, color=C_TEXT, size=11)

def _cell_font():
    return Font(name="Calibri", color=C_TEXT, size=10)

def _header_fill():
    return PatternFill("solid", fgColor=C_HEADER)

def _thin_border():
    s = Side(border_style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def _write_header_row(ws, row, headers, col_widths=None):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _header_font()
        c.fill = _header_fill()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _thin_border()
        if col_widths and i <= len(col_widths):
            _set_col_width(ws, i, col_widths[i - 1])


# ─────────────────────────────────────────────────────────────────────────────
# 1. Bar Inventory Master
# ─────────────────────────────────────────────────────────────────────────────
def make_inventory_master():
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.sheet_properties.tabColor = "A0B4CC"

    headers = ["Product", "Category", "Station", "Size (ml)", "Par", "Level (0-1)", "Unit Cost ($)", "Total Value ($)"]
    col_widths = [28, 14, 20, 10, 8, 12, 14, 14]
    _write_header_row(ws, 1, headers, col_widths)

    # Sample data rows (first 10 as examples; user fills in the rest)
    samples = [
        ("Tito's Handmade Vodka", "Vodka", "Well 1", 750, 1.0, 0.8, 24.99),
        ("Ketel One Vodka", "Vodka", "Well 1", 750, 1.0, 0.6, 26.99),
        ("Tanqueray Gin", "Gin", "Well 1", 750, 1.0, 0.9, 22.99),
        ("Bacardi Superior Rum", "Rum", "Well 1", 750, 1.0, 0.5, 16.99),
        ("Jose Cuervo Gold", "Tequila", "Well 1", 750, 1.0, 0.7, 18.99),
        ("Jack Daniel's Tennessee", "Whiskey", "Well 2", 750, 1.0, 0.4, 22.99),
        ("Maker's Mark Bourbon", "Bourbon", "Well 2", 750, 1.0, 0.8, 27.99),
        ("Jameson Irish Whiskey", "Whiskey", "Well 2", 750, 1.0, 0.6, 23.99),
        ("Hendrick's Gin", "Gin", "Back Bar", 750, 0.5, 1.0, 34.99),
        ("Patron Silver Tequila", "Tequila", "Back Bar", 750, 0.5, 0.9, 44.99),
    ]

    for r, row in enumerate(samples, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c_idx, value=val)
            cell.font = _cell_font()
            cell.border = _thin_border()
            if c_idx == 4:
                cell.number_format = "#,##0"
            if c_idx in (5, 6):
                cell.number_format = "0.0"
            if c_idx == 7:
                cell.number_format = '"$"#,##0.00'
        # Total Value formula
        tf = ws.cell(row=r, column=8,
                     value=f"=G{r}*F{r}")
        tf.font = _cell_font()
        tf.border = _thin_border()
        tf.number_format = '"$"#,##0.00'

    # 10 blank rows for user input
    for r in range(12, 22):
        for c_idx in range(1, 9):
            cell = ws.cell(row=r, column=c_idx, value="")
            cell.font = _cell_font()
            cell.border = _thin_border()
        tf = ws.cell(row=r, column=8, value=f"=IF(G{r}=\"\",\"\",G{r}*F{r})")
        tf.font = _cell_font()
        tf.border = _thin_border()
        tf.number_format = '"$"#,##0.00'

    # Totals row
    last_data = 21
    tot_row = last_data + 1
    tc = ws.cell(row=tot_row, column=1, value="TOTAL")
    tc.font = Font(name="Calibri", bold=True, color=C_TEXT, size=11)
    tc.fill = _header_fill()
    tc.border = _thin_border()
    for c_idx in range(2, 8):
        cell = ws.cell(row=tot_row, column=c_idx, value="")
        cell.fill = _header_fill()
        cell.border = _thin_border()
    tv = ws.cell(row=tot_row, column=8,
                 value=f"=SUM(H2:H{last_data})")
    tv.font = Font(name="Calibri", bold=True, color=C_TEXT, size=11)
    tv.fill = _header_fill()
    tv.border = _thin_border()
    tv.number_format = '"$"#,##0.00'

    # Conditional formatting: red if Level < Par, green if Level >= Par
    red_fill = PatternFill(start_color="FFFDE8E8", end_color="FFFDE8E8", fill_type="solid")
    green_fill = PatternFill(start_color="FFE8F5F0", end_color="FFE8F5F0", fill_type="solid")

    ws.conditional_formatting.add(
        f"F2:F{last_data}",
        FormulaRule(formula=["F2<E2"], fill=red_fill)
    )
    ws.conditional_formatting.add(
        f"F2:F{last_data}",
        FormulaRule(formula=["F2>=E2"], fill=green_fill)
    )

    # Freeze top row
    ws.freeze_panes = "A2"

    wb.save(os.path.join(OUT_DIR, "Bar-Inventory-Master.xlsx"))
    print("  ✓  Bar-Inventory-Master.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 2. Quick Count Sheet
# ─────────────────────────────────────────────────────────────────────────────
def make_quick_count():
    wb = Workbook()
    ws = wb.active
    ws.title = "Quick Count"

    headers = ["Product", "Station", "Last Count", "Count (fill in)", "Variance"]
    col_widths = [30, 20, 14, 16, 12]
    _write_header_row(ws, 1, headers, col_widths)

    # Group of sample products by station
    stations = [
        ("Well 1", [
            ("Tito's Handmade Vodka", 0.8),
            ("Ketel One Vodka", 0.6),
            ("Tanqueray Gin", 0.9),
            ("Bacardi Superior Rum", 0.5),
            ("Jose Cuervo Gold", 0.7),
        ]),
        ("Well 2", [
            ("Jack Daniel's Tennessee", 0.4),
            ("Maker's Mark Bourbon", 0.8),
            ("Jameson Irish Whiskey", 0.6),
            ("Crown Royal Whisky", 0.3),
        ]),
        ("Back Bar", [
            ("Hendrick's Gin", 1.0),
            ("Patron Silver", 0.9),
            ("Don Julio Blanco", 0.7),
            ("Hennessy VS", 0.5),
            ("Woodford Reserve", 0.8),
        ]),
    ]

    row = 2
    for station_name, products in stations:
        # Station header row
        sc = ws.cell(row=row, column=1, value=station_name.upper())
        sc.font = Font(name="Calibri", bold=True, color=C_COPPER, size=11)
        sc.fill = PatternFill("solid", fgColor="FF111822")
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FF111822")
            ws.cell(row=row, column=c).border = _thin_border()
        ws.cell(row=row, column=1).font = Font(name="Calibri", bold=True, color=C_COPPER, size=11)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        for product, last_count in products:
            ws.cell(row=row, column=1, value=product).font = _cell_font()
            ws.cell(row=row, column=2, value=station_name).font = _cell_font()
            lc = ws.cell(row=row, column=3, value=last_count)
            lc.font = _cell_font()
            lc.number_format = "0.0"
            ws.cell(row=row, column=4, value="").font = _cell_font()  # user fills in
            var = ws.cell(row=row, column=5, value=f"=IF(D{row}=\"\",\"\",D{row}-C{row})")
            var.font = _cell_font()
            var.number_format = "0.0"
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = _thin_border()
            row += 1

        row += 1  # blank spacer row

    # Print settings
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:1"
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.freeze_panes = "A2"

    wb.save(os.path.join(OUT_DIR, "Quick-Count-Sheet.xlsx"))
    print("  ✓  Quick-Count-Sheet.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 3. Variance Calculator
# ─────────────────────────────────────────────────────────────────────────────
def make_variance_calculator():
    wb = Workbook()

    # ── Tab 1: Count Entry ──
    ws1 = wb.active
    ws1.title = "Count Entry"
    headers = ["Product", "Previous Level", "Current Level", "Variance", "Notes"]
    _write_header_row(ws1, 1, headers, [28, 16, 16, 12, 24])
    for r in range(2, 22):
        ws1.cell(row=r, column=1, value="").font = _cell_font()
        ws1.cell(row=r, column=2, value="").font = _cell_font()
        ws1.cell(row=r, column=3, value="").font = _cell_font()
        var = ws1.cell(row=r, column=4,
                       value=f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",C{r}-B{r})")
        var.font = _cell_font()
        var.number_format = "0.0"
        ws1.cell(row=r, column=5, value="").font = _cell_font()
        for c in range(1, 6):
            ws1.cell(row=r, column=c).border = _thin_border()

    # ── Tab 2: Invoice Entry ──
    ws2 = wb.create_sheet("Invoice Entry")
    headers2 = ["Product", "Invoice Date", "Size", "Qty Received", "Unit Cost ($)", "Total ($)"]
    _write_header_row(ws2, 1, headers2, [28, 14, 10, 14, 14, 12])
    for r in range(2, 22):
        for c in range(1, 6):
            ws2.cell(row=r, column=c, value="").font = _cell_font()
            ws2.cell(row=r, column=c).border = _thin_border()
        tot = ws2.cell(row=r, column=6,
                       value=f"=IF(OR(D{r}=\"\",E{r}=\"\"),\"\",D{r}*E{r})")
        tot.font = _cell_font()
        tot.number_format = '"$"#,##0.00'
        tot.border = _thin_border()

    # ── Tab 3: Reconciliation ──
    ws3 = wb.create_sheet("Reconciliation")
    headers3 = ["Product", "Prev Level", "Invoiced Qty", "Current Level", "Expected Level", "Actual Usage", "Variance $"]
    _write_header_row(ws3, 1, headers3, [28, 12, 14, 14, 16, 14, 12])
    note = ws3.cell(row=2, column=1,
                    value="Usage = Prev Level + Invoiced Qty - Current Level")
    note.font = Font(name="Calibri", italic=True, color=C_MUTED, size=9)
    for r in range(3, 23):
        for c in range(1, 5):
            ws3.cell(row=r, column=c, value="").font = _cell_font()
            ws3.cell(row=r, column=c).border = _thin_border()
        exp = ws3.cell(row=r, column=5,
                       value=f"=IF(OR(B{r}=\"\",C{r}=\"\"),\"\",B{r}+C{r})")
        exp.font = _cell_font()
        exp.number_format = "0.0"
        exp.border = _thin_border()
        usage = ws3.cell(row=r, column=6,
                         value=f"=IF(OR(B{r}=\"\",D{r}=\"\"),\"\",B{r}+C{r}-D{r})")
        usage.font = _cell_font()
        usage.number_format = "0.0"
        usage.border = _thin_border()
        var = ws3.cell(row=r, column=7, value="")
        var.font = _cell_font()
        var.number_format = '"$"#,##0.00'
        var.border = _thin_border()

    # ── Tab 4: Cost Per Oz ──
    ws4 = wb.create_sheet("Cost Per Oz")
    headers4 = ["Product", "Category", "Size (ml)", "Unit Cost ($)", "Oz per Bottle", "Cost Per Oz ($)", "Std Pour (oz)", "Pour Cost ($)"]
    _write_header_row(ws4, 1, headers4, [28, 14, 10, 14, 14, 14, 12, 12])
    note4 = ws4.cell(row=2, column=1, value="1 fl oz = 29.5735 ml")
    note4.font = Font(name="Calibri", italic=True, color=C_MUTED, size=9)

    sample_cost_data = [
        ("Tito's Handmade Vodka", "Vodka", 750, 24.99, 1.5),
        ("Ketel One Vodka", "Vodka", 750, 26.99, 1.5),
        ("Hendrick's Gin", "Gin", 750, 34.99, 1.5),
        ("Patron Silver", "Tequila", 750, 44.99, 1.5),
        ("Maker's Mark", "Bourbon", 750, 27.99, 1.5),
        ("Jameson", "Whiskey", 750, 23.99, 1.5),
        ("Hennessy VS", "Cognac", 750, 34.99, 1.0),
        ("Bacardi Superior", "Rum", 750, 16.99, 1.5),
    ]
    for r, (prod, cat, size_ml, cost, pour) in enumerate(sample_cost_data, 3):
        ws4.cell(row=r, column=1, value=prod).font = _cell_font()
        ws4.cell(row=r, column=2, value=cat).font = _cell_font()
        ws4.cell(row=r, column=3, value=size_ml).font = _cell_font()
        pc = ws4.cell(row=r, column=4, value=cost)
        pc.font = _cell_font()
        pc.number_format = '"$"#,##0.00'
        oz = ws4.cell(row=r, column=5, value=f"=C{r}/29.5735")
        oz.font = _cell_font()
        oz.number_format = "0.00"
        cpo = ws4.cell(row=r, column=6, value=f"=IF(C{r}=0,\"\",D{r}/(C{r}/29.5735))")
        cpo.font = _cell_font()
        cpo.number_format = '"$"#,##0.000'
        ws4.cell(row=r, column=7, value=pour).font = _cell_font()
        pour_cost = ws4.cell(row=r, column=8, value=f"=F{r}*G{r}")
        pour_cost.font = _cell_font()
        pour_cost.number_format = '"$"#,##0.00'
        for c in range(1, 9):
            ws4.cell(row=r, column=c).border = _thin_border()

    # Blank rows for user input
    for r in range(len(sample_cost_data) + 3, len(sample_cost_data) + 13):
        for c in range(1, 5):
            ws4.cell(row=r, column=c, value="").font = _cell_font()
            ws4.cell(row=r, column=c).border = _thin_border()
        oz = ws4.cell(row=r, column=5, value=f"=IF(C{r}=\"\",\"\",C{r}/29.5735)")
        oz.font = _cell_font()
        oz.number_format = "0.00"
        oz.border = _thin_border()
        cpo = ws4.cell(row=r, column=6, value=f"=IF(OR(C{r}=\"\",D{r}=\"\"),\"\",D{r}/(C{r}/29.5735))")
        cpo.font = _cell_font()
        cpo.number_format = '"$"#,##0.000'
        cpo.border = _thin_border()
        ws4.cell(row=r, column=7, value="").border = _thin_border()
        pc = ws4.cell(row=r, column=8, value=f"=IF(OR(F{r}=\"\",G{r}=\"\"),\"\",F{r}*G{r})")
        pc.font = _cell_font()
        pc.number_format = '"$"#,##0.00'
        pc.border = _thin_border()

    wb.save(os.path.join(OUT_DIR, "Variance-Calculator.xlsx"))
    print("  ✓  Variance-Calculator.xlsx")


# ─────────────────────────────────────────────────────────────────────────────
# 4. Product Database
# ─────────────────────────────────────────────────────────────────────────────
PRODUCTS = [
    # (Name, Category, Size_ml, Pour_oz, Cost_low, Cost_high)
    # Vodka
    ("Tito's Handmade Vodka", "Vodka", 750, 1.5, 20, 26),
    ("Tito's Handmade Vodka", "Vodka", 1750, 1.5, 44, 52),
    ("Ketel One Vodka", "Vodka", 750, 1.5, 24, 30),
    ("Ketel One Vodka", "Vodka", 1750, 1.5, 52, 60),
    ("Grey Goose Vodka", "Vodka", 750, 1.5, 34, 42),
    ("Grey Goose Vodka", "Vodka", 1000, 1.5, 42, 50),
    ("Belvedere Vodka", "Vodka", 750, 1.5, 36, 44),
    ("Absolut Vodka", "Vodka", 750, 1.5, 18, 24),
    ("Absolut Vodka", "Vodka", 1750, 1.5, 36, 44),
    ("Stolichnaya Vodka", "Vodka", 750, 1.5, 16, 22),
    ("Smirnoff No. 21 Vodka", "Vodka", 750, 1.5, 12, 18),
    ("Smirnoff No. 21 Vodka", "Vodka", 1750, 1.5, 24, 32),
    ("Skyy Vodka", "Vodka", 750, 1.5, 14, 20),
    ("Skyy Vodka", "Vodka", 1750, 1.5, 28, 36),
    ("Ciroc Vodka", "Vodka", 750, 1.5, 30, 38),
    ("Pinnacle Vodka", "Vodka", 750, 1.5, 10, 16),
    ("Pinnacle Vodka", "Vodka", 1750, 1.5, 20, 28),
    ("New Amsterdam Vodka", "Vodka", 750, 1.5, 10, 16),
    ("New Amsterdam Vodka", "Vodka", 1750, 1.5, 20, 28),
    ("Burnett's Vodka", "Vodka", 750, 1.5, 8, 14),
    ("Wheatley Vodka", "Vodka", 750, 1.5, 14, 20),
    ("Deep Eddy Vodka", "Vodka", 750, 1.5, 16, 22),
    ("Hangar 1 Vodka", "Vodka", 750, 1.5, 28, 36),
    ("Prairie Organic Vodka", "Vodka", 750, 1.5, 20, 26),
    ("UV Vodka", "Vodka", 750, 1.5, 10, 16),
    # Gin
    ("Tanqueray Gin", "Gin", 750, 1.5, 20, 26),
    ("Tanqueray Gin", "Gin", 1000, 1.5, 26, 32),
    ("Hendrick's Gin", "Gin", 750, 1.5, 30, 38),
    ("Bombay Sapphire Gin", "Gin", 750, 1.5, 22, 28),
    ("Bombay Sapphire Gin", "Gin", 1750, 1.5, 46, 56),
    ("Beefeater Gin", "Gin", 750, 1.5, 18, 24),
    ("Beefeater Gin", "Gin", 1750, 1.5, 36, 44),
    ("Gordon's Gin", "Gin", 750, 1.5, 14, 20),
    ("Seagram's Extra Dry Gin", "Gin", 750, 1.5, 10, 16),
    ("Seagram's Extra Dry Gin", "Gin", 1750, 1.5, 20, 28),
    ("Aviation American Gin", "Gin", 750, 1.5, 26, 34),
    ("The Botanist Gin", "Gin", 750, 1.5, 34, 42),
    ("St. George Terroir Gin", "Gin", 750, 1.5, 32, 40),
    ("Monkey 47 Gin", "Gin", 500, 1.5, 42, 52),
    ("Nolet's Silver Gin", "Gin", 750, 1.5, 38, 48),
    ("Roku Japanese Gin", "Gin", 750, 1.5, 28, 36),
    ("Empress 1908 Gin", "Gin", 750, 1.5, 32, 40),
    # Rum
    ("Bacardi Superior Rum", "Rum", 750, 1.5, 14, 20),
    ("Bacardi Superior Rum", "Rum", 1750, 1.5, 28, 36),
    ("Captain Morgan Spiced Rum", "Rum", 750, 1.5, 14, 20),
    ("Captain Morgan Spiced Rum", "Rum", 1750, 1.5, 28, 36),
    ("Malibu Coconut Rum", "Rum", 750, 1.5, 14, 20),
    ("Malibu Coconut Rum", "Rum", 1750, 1.5, 28, 36),
    ("Kraken Black Rum", "Rum", 750, 1.5, 18, 24),
    ("Mount Gay Eclipse Rum", "Rum", 750, 1.5, 20, 26),
    ("Goslings Black Seal Rum", "Rum", 750, 1.5, 18, 24),
    ("Myers's Dark Rum", "Rum", 750, 1.5, 16, 22),
    ("Havana Club 3yr Rum", "Rum", 750, 1.5, 16, 22),
    ("Plantation 3 Stars Rum", "Rum", 750, 1.5, 18, 24),
    ("Appleton Estate Rum", "Rum", 750, 1.5, 22, 28),
    ("Ron Zacapa 23yr Rum", "Rum", 750, 1.5, 42, 52),
    ("Flor de Caña 7yr Rum", "Rum", 750, 1.5, 22, 28),
    ("Smith & Cross Rum", "Rum", 750, 1.5, 24, 32),
    ("Wray & Nephew White Rum", "Rum", 750, 1.5, 20, 26),
    # Tequila
    ("Jose Cuervo Especial Gold", "Tequila", 750, 1.5, 16, 22),
    ("Jose Cuervo Especial Gold", "Tequila", 1750, 1.5, 32, 42),
    ("Patron Silver Tequila", "Tequila", 750, 1.5, 40, 50),
    ("Patron Silver Tequila", "Tequila", 1000, 1.5, 52, 64),
    ("Don Julio Blanco", "Tequila", 750, 1.5, 46, 56),
    ("Don Julio Reposado", "Tequila", 750, 1.5, 48, 60),
    ("Don Julio 1942", "Tequila", 750, 1.5, 110, 140),
    ("Casamigos Blanco", "Tequila", 750, 1.5, 42, 52),
    ("Casamigos Reposado", "Tequila", 750, 1.5, 44, 54),
    ("Espolon Blanco Tequila", "Tequila", 750, 1.5, 22, 28),
    ("Espolon Reposado Tequila", "Tequila", 750, 1.5, 24, 30),
    ("Hornitos Plata", "Tequila", 750, 1.5, 20, 26),
    ("Olmeca Altos Plata", "Tequila", 750, 1.5, 22, 28),
    ("Milagro Silver Tequila", "Tequila", 750, 1.5, 22, 28),
    ("El Jimador Silver", "Tequila", 750, 1.5, 16, 22),
    ("1800 Silver Tequila", "Tequila", 750, 1.5, 22, 28),
    ("Herradura Silver", "Tequila", 750, 1.5, 32, 40),
    ("Codigo 1530 Blanco", "Tequila", 750, 1.5, 40, 50),
    ("Clase Azul Plata", "Tequila", 750, 1.5, 90, 110),
    # Mezcal
    ("Del Maguey Vida Mezcal", "Mezcal", 750, 1.0, 38, 48),
    ("Ilegal Mezcal Joven", "Mezcal", 750, 1.0, 32, 40),
    ("El Silencio Espadin Mezcal", "Mezcal", 750, 1.0, 22, 28),
    ("Banhez Mezcal", "Mezcal", 750, 1.0, 28, 36),
    ("Putaendo Mezcal", "Mezcal", 750, 1.0, 36, 46),
    # Bourbon
    ("Maker's Mark Bourbon", "Bourbon", 750, 1.5, 24, 32),
    ("Maker's Mark Bourbon", "Bourbon", 1750, 1.5, 50, 62),
    ("Woodford Reserve Bourbon", "Bourbon", 750, 1.5, 30, 38),
    ("Bulleit Bourbon", "Bourbon", 750, 1.5, 26, 34),
    ("Bulleit Bourbon", "Bourbon", 1750, 1.5, 54, 64),
    ("Buffalo Trace Bourbon", "Bourbon", 750, 1.5, 22, 30),
    ("Knob Creek 9yr Bourbon", "Bourbon", 750, 1.5, 30, 38),
    ("Wild Turkey 101 Bourbon", "Bourbon", 750, 1.5, 22, 28),
    ("Four Roses Bourbon", "Bourbon", 750, 1.5, 22, 28),
    ("Eagle Rare 10yr Bourbon", "Bourbon", 750, 1.5, 30, 40),
    ("Blanton's Single Barrel", "Bourbon", 750, 1.5, 60, 80),
    ("Jim Beam White Label", "Bourbon", 750, 1.5, 14, 20),
    ("Jim Beam White Label", "Bourbon", 1750, 1.5, 28, 38),
    ("Basil Hayden's Bourbon", "Bourbon", 750, 1.5, 38, 48),
    ("Angel's Envy Bourbon", "Bourbon", 750, 1.5, 44, 56),
    ("Pappy Van Winkle 15yr", "Bourbon", 750, 1.5, 200, 300),
    # Whiskey / American
    ("Jack Daniel's Old No. 7", "Whiskey", 750, 1.5, 20, 26),
    ("Jack Daniel's Old No. 7", "Whiskey", 1750, 1.5, 40, 52),
    ("Jack Daniel's Tennessee Honey", "Whiskey", 750, 1.5, 20, 26),
    ("Jameson Irish Whiskey", "Whiskey", 750, 1.5, 20, 26),
    ("Jameson Irish Whiskey", "Whiskey", 1750, 1.5, 40, 52),
    ("Crown Royal Whisky", "Whiskey", 750, 1.5, 24, 30),
    ("Crown Royal Whisky", "Whiskey", 1750, 1.5, 48, 60),
    ("Fireball Cinnamon Whisky", "Whiskey", 750, 1.5, 14, 20),
    ("Fireball Cinnamon Whisky", "Whiskey", 1750, 1.5, 28, 38),
    ("Tullamore D.E.W. Irish Whiskey", "Whiskey", 750, 1.5, 20, 26),
    ("Bushmills Original", "Whiskey", 750, 1.5, 18, 24),
    ("Bulleit Rye", "Rye Whiskey", 750, 1.5, 26, 34),
    ("WhistlePig 10yr Rye", "Rye Whiskey", 750, 1.5, 62, 76),
    ("High West Double Rye", "Rye Whiskey", 750, 1.5, 34, 44),
    ("Rittenhouse Rye 100", "Rye Whiskey", 750, 1.5, 22, 28),
    # Scotch
    ("Johnnie Walker Black Label", "Scotch", 750, 1.5, 30, 40),
    ("Johnnie Walker Red Label", "Scotch", 750, 1.5, 22, 28),
    ("Johnnie Walker Double Black", "Scotch", 750, 1.5, 34, 44),
    ("Dewar's White Label", "Scotch", 750, 1.5, 18, 24),
    ("Chivas Regal 12yr", "Scotch", 750, 1.5, 28, 36),
    ("The Glenlivet 12yr", "Scotch", 750, 1.5, 30, 40),
    ("Glenfiddich 12yr", "Scotch", 750, 1.5, 32, 42),
    ("Macallan 12yr Sherry Oak", "Scotch", 750, 1.5, 56, 70),
    ("Laphroaig 10yr", "Scotch", 750, 1.5, 36, 46),
    ("Lagavulin 16yr", "Scotch", 750, 1.5, 60, 76),
    ("Balvenie 12yr DoubleWood", "Scotch", 750, 1.5, 46, 58),
    ("Highland Park 12yr", "Scotch", 750, 1.5, 34, 44),
    # Brandy / Cognac
    ("Hennessy VS Cognac", "Cognac", 750, 1.0, 32, 42),
    ("Hennessy VS Cognac", "Cognac", 1750, 1.0, 70, 90),
    ("Hennessy VSOP Cognac", "Cognac", 750, 1.0, 52, 64),
    ("Remy Martin VSOP Cognac", "Cognac", 750, 1.0, 48, 60),
    ("Courvoisier VS Cognac", "Cognac", 750, 1.0, 30, 40),
    ("E&J VS Brandy", "Brandy", 750, 1.0, 14, 20),
    ("E&J VS Brandy", "Brandy", 1750, 1.0, 28, 38),
    ("Paul Masson Grande Amber VSOP", "Brandy", 750, 1.0, 12, 18),
    ("Korbel California Brandy", "Brandy", 750, 1.0, 12, 18),
    # Liqueur
    ("Kahlúa Coffee Liqueur", "Liqueur", 750, 1.0, 18, 24),
    ("Baileys Irish Cream", "Liqueur", 750, 1.0, 22, 28),
    ("Amaretto Disaronno", "Liqueur", 750, 1.0, 24, 30),
    ("Cointreau Triple Sec", "Liqueur", 750, 1.0, 32, 40),
    ("Grand Marnier", "Liqueur", 750, 1.0, 32, 40),
    ("Chambord Black Raspberry", "Liqueur", 750, 0.5, 28, 36),
    ("Midori Melon Liqueur", "Liqueur", 750, 0.75, 18, 24),
    ("Frangelico Hazelnut Liqueur", "Liqueur", 750, 0.75, 22, 28),
    ("St-Germain Elderflower", "Liqueur", 750, 0.5, 26, 34),
    ("Aperol", "Liqueur", 750, 1.0, 18, 24),
    ("Campari", "Liqueur", 750, 1.0, 20, 26),
    ("Chartreuse Green", "Liqueur", 750, 0.5, 44, 56),
    ("Chartreuse Yellow", "Liqueur", 750, 0.5, 38, 48),
    ("Benedictine", "Liqueur", 750, 0.5, 28, 36),
    ("Drambuie", "Liqueur", 750, 0.75, 28, 36),
    ("Galliano L'Autentico", "Liqueur", 750, 0.5, 24, 32),
    ("Triple Sec (Generic)", "Liqueur", 750, 1.0, 8, 14),
    ("Blue Curaçao", "Liqueur", 750, 0.75, 12, 18),
    ("Peach Schnapps", "Liqueur", 750, 1.0, 10, 16),
    ("Sloe Gin", "Liqueur", 750, 1.0, 16, 22),
    ("Limoncello", "Liqueur", 750, 1.0, 18, 24),
    ("Tuaca Vanilla Citrus", "Liqueur", 750, 0.75, 18, 24),
    # Amaro / Bitter
    ("Fernet-Branca", "Amaro", 750, 1.0, 28, 36),
    ("Fernet-Branca", "Amaro", 1000, 1.0, 36, 46),
    ("Cynar Artichoke Amaro", "Amaro", 750, 1.0, 18, 24),
    ("Averna Amaro", "Amaro", 750, 1.0, 22, 28),
    ("Montenegro Amaro", "Amaro", 750, 1.0, 22, 28),
    ("Nonino Amaro", "Amaro", 750, 1.0, 38, 48),
    ("Luxardo Maraschino", "Liqueur", 750, 0.25, 22, 28),
    ("Luxardo Amaretto", "Liqueur", 750, 1.0, 18, 24),
    ("Angostura Aromatic Bitters", "Bitters", 200, 0.0, 7, 11),
    ("Angostura Orange Bitters", "Bitters", 200, 0.0, 7, 11),
    ("Peychaud's Bitters", "Bitters", 150, 0.0, 7, 11),
    ("Fee Brothers Old Fashioned", "Bitters", 150, 0.0, 6, 10),
    ("Regans' Orange Bitters", "Bitters", 150, 0.0, 8, 12),
    # Vermouth / Fortified
    ("Martini & Rossi Sweet Vermouth", "Vermouth", 750, 1.0, 10, 14),
    ("Martini & Rossi Dry Vermouth", "Vermouth", 750, 1.0, 10, 14),
    ("Noilly Prat Dry Vermouth", "Vermouth", 750, 1.0, 12, 16),
    ("Dolin Dry Vermouth", "Vermouth", 750, 1.0, 12, 16),
    ("Dolin Sweet Vermouth", "Vermouth", 750, 1.0, 12, 16),
    ("Carpano Antica Formula", "Vermouth", 750, 1.0, 22, 28),
    ("Punt e Mes", "Vermouth", 750, 1.0, 18, 24),
    ("Lillet Blanc", "Fortified Wine", 750, 3.0, 16, 22),
    ("Aperol Spritz Ready-to-Drink", "Fortified Wine", 750, 3.0, 14, 20),
    # Wine (common bar pours)
    ("House Red Wine (Bottle)", "Wine", 750, 5.0, 8, 14),
    ("House White Wine (Bottle)", "Wine", 750, 5.0, 8, 14),
    ("House Rosé (Bottle)", "Wine", 750, 5.0, 10, 16),
    ("Meiomi Pinot Noir", "Wine", 750, 5.0, 14, 20),
    ("Kim Crawford Sauvignon Blanc", "Wine", 750, 5.0, 14, 20),
    ("La Marca Prosecco", "Sparkling Wine", 750, 5.0, 14, 20),
    # Mixers
    ("Club Soda (Liter)", "Mixer", 1000, 4.0, 1, 2),
    ("Tonic Water (Liter)", "Mixer", 1000, 4.0, 1, 3),
    ("Ginger Beer (Bottle)", "Mixer", 355, 4.0, 1, 2),
    ("Ginger Ale (Liter)", "Mixer", 1000, 4.0, 1, 2),
    ("Cranberry Juice (Liter)", "Mixer", 1000, 4.0, 3, 6),
    ("Orange Juice (Quart)", "Mixer", 946, 4.0, 3, 5),
    ("Pineapple Juice (Can)", "Mixer", 946, 4.0, 2, 4),
    ("Lime Juice (Bottle)", "Mixer", 750, 0.75, 4, 8),
    ("Lemon Juice (Bottle)", "Mixer", 750, 0.75, 4, 8),
    ("Grenadine (Bottle)", "Syrup", 750, 0.5, 6, 10),
    ("Simple Syrup (Bottle)", "Syrup", 750, 0.5, 4, 8),
    ("Agave Nectar (Bottle)", "Syrup", 750, 0.5, 6, 10),
    ("Orgeat Almond Syrup", "Syrup", 750, 0.5, 10, 16),
    ("Honey Syrup (House)", "Syrup", 750, 0.5, 4, 8),
    ("Blue Curacao Syrup", "Syrup", 750, 0.5, 6, 10),
    ("Passion Fruit Syrup", "Syrup", 750, 0.5, 8, 14),
    ("Falernum", "Syrup", 750, 0.5, 14, 20),
    ("Coke (2L)", "Soda", 2000, 6.0, 2, 4),
    ("Diet Coke (2L)", "Soda", 2000, 6.0, 2, 4),
    ("Sprite (2L)", "Soda", 2000, 6.0, 2, 4),
    ("Redbull (8.4oz)", "Energy Drink", 250, 8.4, 2, 3),
    ("Monster Energy (16oz)", "Energy Drink", 473, 16.0, 2, 3),
]

def make_product_database():
    wb = Workbook()
    ws = wb.active
    ws.title = "Product Database"

    headers = ["Name", "Category", "Size (ml)", "Std Pour (oz)", "Cost Low ($)", "Cost High ($)", "Cost/Oz Low", "Cost/Oz High"]
    col_widths = [36, 16, 10, 12, 12, 12, 13, 13]
    _write_header_row(ws, 1, headers, col_widths)

    for r, (name, cat, size_ml, pour, cost_low, cost_high) in enumerate(PRODUCTS, 2):
        ws.cell(row=r, column=1, value=name).font = _cell_font()
        ws.cell(row=r, column=2, value=cat).font = _cell_font()
        ws.cell(row=r, column=3, value=size_ml).font = _cell_font()
        ws.cell(row=r, column=4, value=pour).font = _cell_font()
        cl = ws.cell(row=r, column=5, value=cost_low)
        cl.font = _cell_font()
        cl.number_format = '"$"#,##0.00'
        ch = ws.cell(row=r, column=6, value=cost_high)
        ch.font = _cell_font()
        ch.number_format = '"$"#,##0.00'
        # Cost per oz calculations
        cpo_low = ws.cell(row=r, column=7, value=f"=IF(C{r}=0,\"\",E{r}/(C{r}/29.5735))")
        cpo_low.font = _cell_font()
        cpo_low.number_format = '"$"#,##0.000'
        cpo_high = ws.cell(row=r, column=8, value=f"=IF(C{r}=0,\"\",F{r}/(C{r}/29.5735))")
        cpo_high.font = _cell_font()
        cpo_high.number_format = '"$"#,##0.000'
        for c in range(1, 9):
            ws.cell(row=r, column=c).border = _thin_border()

    # Auto-filter + freeze
    ws.auto_filter.ref = f"A1:H{len(PRODUCTS) + 1}"
    ws.freeze_panes = "A2"

    wb.save(os.path.join(OUT_DIR, "Product-Database.xlsx"))
    print(f"  ✓  Product-Database.xlsx  ({len(PRODUCTS)} products)")


if __name__ == "__main__":
    print(f"\nGenerating XLSX templates → {OUT_DIR}\n")
    make_inventory_master()
    make_quick_count()
    make_variance_calculator()
    make_product_database()
    print("\nDone.\n")
