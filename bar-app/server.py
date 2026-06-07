"""
Bar Inventory Server — Open Source Barware
Flask app serving dashboard.html with JSON API and file-based persistence.
"""

import io, json, os, re, ssl, sys, uuid, urllib.request, urllib.error
from datetime import datetime, timezone
from flask import Flask, request, jsonify, send_from_directory, send_file

# ─────────────────────────────────────────────
# PATHS
# ─────────────────────────────────────────────
_DIR = os.path.dirname(os.path.abspath(__file__))
_BAR_FILE = os.path.join(_DIR, "bar_data.json")
_COUNT_FILE = os.path.join(_DIR, "count_history.json")
_CONFIG_FILE = os.path.join(_DIR, "config.json")
_INVOICE_FILE = os.path.join(_DIR, "invoice_history.json")

# ─────────────────────────────────────────────
# SSL — same macOS fix as the OVLP pop
# ─────────────────────────────────────────────
def _ssl():
    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except ImportError:
        ctx = ssl.SSLContext(ssl.PROTOCOL_TLS_CLIENT)
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    try:
        ctx.maximum_version = ssl.TLSVersion.TLSv1_2
    except (AttributeError, ValueError):
        pass
    try:
        ctx.set_ciphers("DEFAULT:@SECLEVEL=1")
    except ssl.SSLError:
        pass
    return ctx

# ─────────────────────────────────────────────
# FLASK APP
# ─────────────────────────────────────────────
app = Flask(__name__, static_folder=".")


@app.after_request
def _cors(resp):
    resp.headers["Access-Control-Allow-Origin"] = "*"
    resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
    resp.headers["Access-Control-Allow-Methods"] = "GET, POST, PUT, DELETE, OPTIONS"
    return resp


# ─────────────────────────────────────────────
# PERSISTENCE HELPERS
# ─────────────────────────────────────────────
def _now():
    return datetime.now(timezone.utc).isoformat()


def _uid(prefix=""):
    return f"{prefix}{uuid.uuid4().hex[:8]}"


def _load_bar():
    """Load bar_data.json or return empty scaffold."""
    if os.path.exists(_BAR_FILE):
        try:
            with open(_BAR_FILE) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"bar_name": "", "created": "", "stations": []}


def _save_bar(data):
    tmp = _BAR_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, _BAR_FILE)


def _load_counts():
    if os.path.exists(_COUNT_FILE):
        try:
            with open(_COUNT_FILE) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return []


def _save_counts(data):
    tmp = _COUNT_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, _COUNT_FILE)


def _find_station(bar, station_id):
    """Return (index, station_dict) or (None, None)."""
    for i, s in enumerate(bar.get("stations", [])):
        if s["id"] == station_id:
            return i, s
    return None, None


def _find_bottle(bar, bottle_id):
    """Return (station_dict, bottle_index, bottle_dict) or (None, None, None)."""
    for s in bar.get("stations", []):
        for j, b in enumerate(s.get("bottles", [])):
            if b["id"] == bottle_id:
                return s, j, b
    return None, None, None


def _all_bottles(bar):
    """Flat list of all bottles across all stations."""
    out = []
    for s in bar.get("stations", []):
        for b in s.get("bottles", []):
            out.append(b)
    return out


def _load_config():
    if os.path.exists(_CONFIG_FILE):
        try:
            with open(_CONFIG_FILE) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return {"api_base_url": "", "api_key": "", "model": ""}


def _save_config(data):
    tmp = _CONFIG_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, _CONFIG_FILE)


def _load_invoices():
    if os.path.exists(_INVOICE_FILE):
        try:
            with open(_INVOICE_FILE) as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            pass
    return []


def _save_invoices(data):
    tmp = _INVOICE_FILE + ".tmp"
    with open(tmp, "w") as f:
        json.dump(data, f, indent=2)
    os.replace(tmp, _INVOICE_FILE)


def _mask_key(key):
    """Show last 4 chars of API key, mask the rest."""
    if not key:
        return ""
    if len(key) <= 4:
        return "*" * len(key)
    return "*" * (len(key) - 4) + key[-4:]


def _call_ai(messages, config):
    """Call any OpenAI-compatible endpoint. Returns content string or raises."""
    base_url = config.get("api_base_url", "").rstrip("/")
    if not base_url:
        raise ValueError("api_base_url not configured")
    url = base_url + "/chat/completions"
    payload = {
        "model": config.get("model", "gpt-4o"),
        "messages": messages,
        "max_tokens": 4096,
    }
    data = json.dumps(payload).encode()
    headers = {
        "Authorization": f"Bearer {config.get('api_key', '')}",
        "Content-Type": "application/json",
    }
    req = urllib.request.Request(url, data=data, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=60) as r:
            body = json.loads(r.read())
            return body["choices"][0]["message"]["content"]
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"AI API error {e.code}: {err_body[:400]}")


# ─────────────────────────────────────────────
# PROMPT TEMPLATES
# ─────────────────────────────────────────────
_PROMPT_SETUP = """\
You are helping set up a bar inventory system.

Parse this bar walk transcription and return ONLY a JSON object with this exact structure — no explanation, no markdown, just the JSON:
{{"stations": [{{"name": "Well 1", "type": "well", "bottles": [{{"name": "Tito's Vodka", "category": "vodka"}}]}}]}}

Station types must be one of: well, backbar, storage, walkin
Categories must be one of: vodka, gin, rum, tequila, mezcal, whiskey, bourbon, scotch, brandy, cognac, liqueur, amaro, wine, beer, mixer, syrup, other

Transcription:
{text}"""

_PROMPT_COUNT = """\
You are parsing a weekly bar inventory count.

Known products (bottle_id: name):
{product_list}

Parse this count transcription and return ONLY a JSON object — no explanation, no markdown, just the JSON:
{{"matched": [{{"bottle_id": "bot-xxx", "bottle_name": "Tito's Vodka", "level": 0.7, "confidence": "high"}}], "unmatched": [{{"text": "...", "reason": "..."}}]}}

Level rules (must be 0.0 to 1.0 in tenths):
- "nine-tenths" or "nine tenths" = 0.9
- "half" = 0.5
- "point three" = 0.3
- "full" or "new" = 1.0
- "empty" = 0.0
- "three quarters" = 0.75
- A digit 1–9 alone likely means tenths (e.g. "Tito's nine" = 0.9)

Transcription:
{text}"""

_PROMPT_INVOICE = """\
You are extracting line items from a liquor/beverage supply invoice image.

Return ONLY a JSON object — no explanation, no markdown, just the JSON:
{{"items": [{{"name": "Tito's Handmade Vodka 1L", "size": "1L", "qty": 6, "unit_cost": 22.99, "total": 137.94}}]}}

Rules:
- Extract every product line item visible
- If qty or cost is unclear or missing, use null
- size should be the bottle size (e.g. "750ml", "1L", "1.75L")
- Do not include subtotals, taxes, or delivery fees as line items"""


# ─────────────────────────────────────────────
# ROUTES — Dashboard
# ─────────────────────────────────────────────
@app.route("/")
def index():
    return send_from_directory(_DIR, "dashboard.html")


# ─────────────────────────────────────────────
# ROUTES — Bar Setup & Config
# ─────────────────────────────────────────────
@app.route("/api/bar", methods=["GET"])
def get_bar():
    return jsonify(_load_bar())


@app.route("/api/bar/setup", methods=["POST"])
def setup_bar():
    """Initial setup — bar name and optional starter stations."""
    body = request.get_json(force=True)
    bar = _load_bar()
    bar["bar_name"] = body.get("bar_name", bar.get("bar_name", "My Bar"))
    if not bar["created"]:
        bar["created"] = _now()

    # Accept optional stations array in setup
    for s in body.get("stations", []):
        station = {
            "id": s.get("id") or _uid("stn-"),
            "name": s.get("name", "Station"),
            "type": s.get("type", "well"),
            "position": s.get("position", len(bar["stations"])),
            "bottles": [],
        }
        bar["stations"].append(station)

    _save_bar(bar)
    print(f"[setup] bar configured: {bar['bar_name']}, {len(bar['stations'])} stations")
    return jsonify(bar), 200


@app.route("/api/bar/station", methods=["POST"])
def add_station():
    body = request.get_json(force=True)
    bar = _load_bar()
    station = {
        "id": body.get("id") or _uid("stn-"),
        "name": body.get("name", "Station"),
        "type": body.get("type", "well"),
        "position": body.get("position", len(bar["stations"])),
        "bottles": [],
    }
    bar["stations"].append(station)
    _save_bar(bar)
    print(f"[station] added: {station['name']} ({station['id']})")
    return jsonify(station), 201


@app.route("/api/bar/station/<station_id>", methods=["PUT"])
def update_station(station_id):
    body = request.get_json(force=True)
    bar = _load_bar()
    idx, station = _find_station(bar, station_id)
    if station is None:
        return jsonify({"error": "station not found"}), 404
    for key in ("name", "type", "position"):
        if key in body:
            station[key] = body[key]
    bar["stations"][idx] = station
    _save_bar(bar)
    print(f"[station] updated: {station['name']} ({station_id})")
    return jsonify(station), 200


@app.route("/api/bar/station/<station_id>", methods=["DELETE"])
def delete_station(station_id):
    bar = _load_bar()
    idx, station = _find_station(bar, station_id)
    if station is None:
        return jsonify({"error": "station not found"}), 404
    bar["stations"].pop(idx)
    _save_bar(bar)
    print(f"[station] deleted: {station_id}")
    return jsonify({"deleted": station_id}), 200


# ─────────────────────────────────────────────
# ROUTES — Bottles
# ─────────────────────────────────────────────
@app.route("/api/bottle", methods=["POST"])
def add_bottle():
    body = request.get_json(force=True)
    bar = _load_bar()
    sid = body.get("station_id")
    if not sid:
        return jsonify({"error": "station_id required"}), 400
    idx, station = _find_station(bar, sid)
    if station is None:
        return jsonify({"error": "station not found"}), 404

    bottle = {
        "id": _uid("bot-"),
        "name": body.get("name", "Unknown"),
        "category": body.get("category", "spirits"),
        "size": body.get("size", "750ml"),
        "par_level": float(body.get("par_level", 1.0)),
        "current_level": float(body.get("current_level", 1.0)),
        "cost": float(body.get("cost", 0.0)),
        "last_counted": "",
    }
    station["bottles"].append(bottle)
    _save_bar(bar)
    print(f"[bottle] added: {bottle['name']} -> {station['name']}")
    return jsonify(bottle), 201


@app.route("/api/bottle/<bottle_id>", methods=["PUT"])
def update_bottle(bottle_id):
    body = request.get_json(force=True)
    bar = _load_bar()
    station, bidx, bottle = _find_bottle(bar, bottle_id)
    if bottle is None:
        return jsonify({"error": "bottle not found"}), 404
    for key in ("name", "category", "size", "par_level", "current_level", "cost"):
        if key in body:
            if key in ("par_level", "current_level", "cost"):
                bottle[key] = float(body[key])
            else:
                bottle[key] = body[key]
    if "current_level" in body:
        bottle["last_counted"] = _now()
    station["bottles"][bidx] = bottle
    _save_bar(bar)
    print(f"[bottle] updated: {bottle['name']} ({bottle_id})")
    return jsonify(bottle), 200


@app.route("/api/bottle/<bottle_id>", methods=["DELETE"])
def delete_bottle(bottle_id):
    bar = _load_bar()
    station, bidx, bottle = _find_bottle(bar, bottle_id)
    if bottle is None:
        return jsonify({"error": "bottle not found"}), 404
    station["bottles"].pop(bidx)
    _save_bar(bar)
    print(f"[bottle] deleted: {bottle_id}")
    return jsonify({"deleted": bottle_id}), 200


@app.route("/api/bottles/bulk", methods=["POST"])
def bulk_add_bottles():
    """Add multiple bottles at once (setup wizard). Body: {station_id, bottles: [...]}"""
    body = request.get_json(force=True)
    bar = _load_bar()
    sid = body.get("station_id")
    if not sid:
        return jsonify({"error": "station_id required"}), 400
    idx, station = _find_station(bar, sid)
    if station is None:
        return jsonify({"error": "station not found"}), 404

    added = []
    for b in body.get("bottles", []):
        bottle = {
            "id": _uid("bot-"),
            "name": b.get("name", "Unknown"),
            "category": b.get("category", "spirits"),
            "size": b.get("size", "750ml"),
            "par_level": float(b.get("par_level", 1.0)),
            "current_level": float(b.get("current_level", 1.0)),
            "cost": float(b.get("cost", 0.0)),
            "last_counted": "",
        }
        station["bottles"].append(bottle)
        added.append(bottle)

    _save_bar(bar)
    print(f"[bulk] added {len(added)} bottles to {station['name']}")
    return jsonify({"added": len(added), "bottles": added}), 201


# ─────────────────────────────────────────────
# ROUTES — Counting
# ─────────────────────────────────────────────
@app.route("/api/count", methods=["POST"])
def save_count():
    """Save a count session. Body: {entries: [{bottle_id, level, notes}, ...]}"""
    body = request.get_json(force=True)
    entries = body.get("entries", [])
    if not entries:
        return jsonify({"error": "entries required"}), 400

    bar = _load_bar()

    # Update current_level on each bottle
    below_par = 0
    for entry in entries:
        station, bidx, bottle = _find_bottle(bar, entry.get("bottle_id", ""))
        if bottle:
            bottle["current_level"] = float(entry.get("level", bottle["current_level"]))
            bottle["last_counted"] = _now()
            station["bottles"][bidx] = bottle
            if bottle["current_level"] < bottle["par_level"]:
                below_par += 1

    _save_bar(bar)

    # Save count record
    counts = _load_counts()
    record = {
        "id": _uid("cnt-"),
        "date": _now(),
        "entries": entries,
        "summary": {
            "total_counted": len(entries),
            "below_par": below_par,
        },
    }
    counts.append(record)
    _save_counts(counts)
    print(f"[count] saved: {len(entries)} entries, {below_par} below par")
    return jsonify(record), 201


@app.route("/api/counts", methods=["GET"])
def get_counts():
    return jsonify(_load_counts())


@app.route("/api/count/<count_id>", methods=["GET"])
def get_count(count_id):
    for c in _load_counts():
        if c["id"] == count_id:
            return jsonify(c)
    return jsonify({"error": "count not found"}), 404


# ─────────────────────────────────────────────
# ROUTES — Voice/Text Parser
# ─────────────────────────────────────────────

# Level keywords → float value
_LEVEL_MAP = {
    "full": 1.0, "new": 1.0, "sealed": 1.0, "unopened": 1.0,
    "three quarters": 0.75, "three fourths": 0.75, "three-quarters": 0.75,
    "half": 0.5, "half bottle": 0.5,
    "quarter": 0.25, "one quarter": 0.25,
    "empty": 0.0, "done": 0.0, "dead": 0.0, "86": 0.0,
    "eighty-six": 0.0, "eighty six": 0.0,
}

# Common words to strip for fuzzy matching
_STRIP_WORDS = {
    "vodka", "gin", "rum", "tequila", "whiskey", "whisky", "bourbon",
    "scotch", "brandy", "cognac", "liqueur", "wine", "beer", "ale",
    "lager", "ipa", "stout", "porter", "cider", "seltzer", "mixer",
    "bottle", "bottles", "the", "a", "an", "of",
}


def _normalize(name):
    """Lowercase, strip punctuation + common category words, collapse whitespace."""
    import re
    n = re.sub(r"[''`\-.,!?]", "", name.lower().strip())
    words = [w for w in n.split() if w not in _STRIP_WORDS]
    return " ".join(words) if words else n.lower().strip()


def _fuzzy_match(text, bottles):
    """Try to match text against bottle names. Return (bottle, confidence) or (None, None)."""
    text_norm = _normalize(text)
    if not text_norm:
        return None, None

    # Exact match on normalized name
    for b in bottles:
        if _normalize(b["name"]) == text_norm:
            return b, "high"

    # Substring match — text is contained in bottle name or vice versa
    for b in bottles:
        bn = _normalize(b["name"])
        if text_norm in bn or bn in text_norm:
            return b, "medium"

    # Partial word overlap
    text_words = set(text_norm.split())
    best = None
    best_score = 0
    for b in bottles:
        bn_words = set(_normalize(b["name"]).split())
        overlap = len(text_words & bn_words)
        if overlap > best_score and overlap > 0:
            best = b
            best_score = overlap

    if best and best_score >= 1:
        return best, "low"

    return None, None


def _parse_level(text):
    """Extract a level float from text. Return (level, remainder) or (None, text)."""
    t = text.lower().strip()

    # Check multi-word keywords first (longest match first)
    for phrase in sorted(_LEVEL_MAP.keys(), key=len, reverse=True):
        if phrase in t:
            remainder = t.replace(phrase, "", 1).strip()
            return _LEVEL_MAP[phrase], remainder

    # "point X" or "point four" etc.
    import re
    m = re.search(r'point\s+(\d)', t)
    if m:
        val = int(m.group(1)) / 10.0
        remainder = t[:m.start()] + t[m.end():]
        return val, remainder.strip()

    # Word-form "point four", "point seven" etc.
    _word_digits = {
        "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4,
        "five": 5, "six": 6, "seven": 7, "eight": 8, "nine": 9,
    }
    m = re.search(r'point\s+(' + '|'.join(_word_digits.keys()) + r')', t)
    if m:
        val = _word_digits[m.group(1)] / 10.0
        remainder = t[:m.start()] + t[m.end():]
        return val, remainder.strip()

    # Decimal like "0.4" or ".6"
    m = re.search(r'(\d*\.\d+)', t)
    if m:
        val = float(m.group(1))
        if val > 1.0:
            val = val  # could be case count, leave as-is
        remainder = t[:m.start()] + t[m.end():]
        return val, remainder.strip()

    # Storage: "X cases Y bottles" or "X case" or "X bottles"
    m = re.search(r'(\d+)\s*cases?\s*(?:(\d+)\s*bottles?)?', t)
    if m:
        cases = int(m.group(1))
        extra = int(m.group(2)) if m.group(2) else 0
        # A case is typically 12 bottles; represent as full bottle count
        total = cases * 12 + extra
        remainder = t[:m.start()] + t[m.end():]
        return float(total), remainder.strip()

    m = re.search(r'(\d+)\s*bottles?', t)
    if m:
        remainder = t[:m.start()] + t[m.end():]
        return float(int(m.group(1))), remainder.strip()

    # Bare integer or float
    m = re.search(r'\b(\d+)\b', t)
    if m:
        val = int(m.group(1))
        # Single digit 1-9 alone — likely a tenths level (e.g., "3" = 0.3)
        # But could also be a count. Context-dependent; default to tenths if <=9
        remainder = t[:m.start()] + t[m.end():]
        if 1 <= val <= 9:
            return val / 10.0, remainder.strip()
        return float(val), remainder.strip()

    return None, t


@app.route("/api/parse-notes", methods=["POST"])
def parse_notes():
    """Parse voice/text notes into structured bottle entries.

    Body: {text: "Well one titos half stoli point two tanqueray empty"}
    Returns: {matched: [...], unmatched: [...]}
    """
    body = request.get_json(force=True)
    raw = body.get("text", "").strip()
    if not raw:
        return jsonify({"error": "text required"}), 400

    bar = _load_bar()
    all_bottles = _all_bottles(bar)
    station_names = {s["name"].lower(): s for s in bar.get("stations", [])}

    # Split by station markers if present
    import re
    segments = []
    station_pattern = "|".join(re.escape(n) for n in station_names.keys()) if station_names else None

    if station_pattern:
        parts = re.split(f"({station_pattern})", raw.lower())
        current_station = None
        for part in parts:
            part = part.strip()
            if not part:
                continue
            if part in station_names:
                current_station = station_names[part]
            else:
                segments.append((current_station, part))
    else:
        segments = [(None, raw.lower())]

    matched = []
    unmatched = []

    for station_ctx, segment in segments:
        # Determine which bottles to search in
        if station_ctx:
            search_bottles = station_ctx.get("bottles", [])
        else:
            search_bottles = all_bottles

        # Tokenize: try to find bottle-name + level pairs
        # Strategy: walk through words, accumulate until we find a level or bottle match
        words = segment.split()
        i = 0
        while i < len(words):
            # Try increasingly long spans for bottle name matching
            best_match = None
            best_confidence = None
            best_end = i
            best_level = None

            for end in range(i + 1, min(i + 6, len(words) + 1)):
                candidate = " ".join(words[i:end])

                bottle, confidence = _fuzzy_match(candidate, search_bottles)
                if bottle:
                    best_match = bottle
                    best_confidence = confidence
                    best_end = end

                    # Look for level after the name
                    remaining = " ".join(words[end:end + 4])
                    level, _ = _parse_level(remaining)
                    if level is not None:
                        best_level = level
                        # Count how many words the level consumed
                        level_text_len = len(remaining) - len(_.strip()) if _ else len(remaining)
                        # Rough word count consumed by level
                        level_words = len(remaining[:level_text_len].split()) if level_text_len > 0 else 0
                        best_end = end + max(level_words, 1)

            if best_match:
                entry = {
                    "bottle_id": best_match["id"],
                    "bottle_name": best_match["name"],
                    "level": best_level if best_level is not None else best_match.get("current_level", 1.0),
                    "confidence": best_confidence if best_level is not None else "low",
                    "notes": "",
                }
                if best_level is not None and best_confidence == "high":
                    entry["confidence"] = "high"
                matched.append(entry)
                i = best_end
            else:
                # Check if this word alone is a level with no bottle context
                level, remainder = _parse_level(words[i])
                if level is not None:
                    unmatched.append({
                        "text": words[i],
                        "parsed_level": level,
                        "reason": "no bottle match",
                    })
                else:
                    # Accumulate as unmatched text
                    unmatched.append({
                        "text": words[i],
                        "parsed_level": None,
                        "reason": "unrecognized",
                    })
                i += 1

    print(f"[parse] {len(matched)} matched, {len(unmatched)} unmatched from {len(raw)} chars")
    return jsonify({"matched": matched, "unmatched": unmatched}), 200


# ─────────────────────────────────────────────
# ROUTES — Config
# ─────────────────────────────────────────────
@app.route("/api/config", methods=["GET"])
def get_config():
    cfg = _load_config()
    return jsonify({
        "api_base_url": cfg.get("api_base_url", ""),
        "api_key_masked": _mask_key(cfg.get("api_key", "")),
        "model": cfg.get("model", ""),
        "configured": bool(cfg.get("api_base_url") and cfg.get("api_key")),
    })


@app.route("/api/config", methods=["POST"])
def save_config():
    body = request.get_json(force=True)
    cfg = _load_config()
    if "api_base_url" in body:
        cfg["api_base_url"] = body["api_base_url"].strip()
    if "api_key" in body:
        key = body["api_key"].strip()
        # Don't overwrite with masked value the client sent back
        if key and not set(key) <= {"*"}:
            cfg["api_key"] = key
    if "model" in body:
        cfg["model"] = body["model"].strip()
    _save_config(cfg)
    print(f"[config] saved: base_url={cfg.get('api_base_url','')}, model={cfg.get('model','')}")
    return jsonify({
        "api_base_url": cfg.get("api_base_url", ""),
        "api_key_masked": _mask_key(cfg.get("api_key", "")),
        "model": cfg.get("model", ""),
        "configured": bool(cfg.get("api_base_url") and cfg.get("api_key")),
    })


# ─────────────────────────────────────────────
# ROUTES — AI
# ─────────────────────────────────────────────
def _build_prompt(ptype, placeholder=True):
    """Build a copy-paste prompt for the given type. Returns prompt str or None if invalid type."""
    bottles = _all_bottles(_load_bar())
    if ptype == "setup":
        text = "[PASTE YOUR BAR WALK TRANSCRIPTION HERE]" if placeholder else ""
        return _PROMPT_SETUP.replace("{text}", text)
    elif ptype == "count":
        product_list = "\n".join(
            f'{b["id"]}: {b["name"]}' for b in bottles
        ) or "(no products set up yet)"
        text = "[PASTE YOUR COUNT TRANSCRIPTION HERE]" if placeholder else ""
        return _PROMPT_COUNT.replace("{product_list}", product_list).replace("{text}", text)
    elif ptype == "invoice":
        return _PROMPT_INVOICE
    return None


@app.route("/api/ai/prompt", methods=["GET"])
def get_ai_prompt():
    """Return a prompt template the user can copy into any AI manually."""
    ptype = request.args.get("type", "")
    prompt = _build_prompt(ptype)
    if prompt is None:
        return jsonify({"error": "type must be setup, count, or invoice"}), 400
    return jsonify({"prompt": prompt, "type": ptype})


@app.route("/api/ai/call", methods=["POST"])
def ai_call():
    """Call configured AI with the appropriate prompt. Falls back to prompt text if unconfigured."""
    body = request.get_json(force=True)
    ptype = body.get("type", "")
    text = body.get("text", "")
    images = body.get("images", [])  # list of base64 data URIs

    if ptype not in ("setup", "count", "invoice"):
        return jsonify({"error": "type must be setup, count, or invoice"}), 400

    cfg = _load_config()
    if not cfg.get("api_base_url") or not cfg.get("api_key"):
        # Return the prompt so client can use it manually
        return jsonify({
            "error": "AI not configured",
            "prompt": _build_prompt(ptype) or "",
        }), 503

    # Build messages
    bar = _load_bar()
    bottles = _all_bottles(bar)

    if ptype == "setup":
        content = _PROMPT_SETUP.replace("{text}", text)
        messages = [{"role": "user", "content": content}]

    elif ptype == "count":
        product_list = "\n".join(f'{b["id"]}: {b["name"]}' for b in bottles)
        content = _PROMPT_COUNT.replace("{product_list}", product_list).replace("{text}", text)
        messages = [{"role": "user", "content": content}]

    else:  # invoice
        # Build multi-modal message with text + images
        content_parts = []
        for img in images:
            # img is a data URI: "data:image/jpeg;base64,..."
            if img.startswith("data:"):
                mime, b64 = img.split(",", 1)
                mime = mime.replace("data:", "").replace(";base64", "")
            else:
                mime = "image/jpeg"
                b64 = img
            content_parts.append({
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{b64}"},
            })
        content_parts.append({"type": "text", "text": _PROMPT_INVOICE})
        messages = [{"role": "user", "content": content_parts}]

    try:
        result = _call_ai(messages, cfg)
        print(f"[ai] {ptype} call succeeded, response length: {len(result)}")
        return jsonify({"result": result})
    except Exception as e:
        print(f"[ai] {ptype} call failed: {e}")
        return jsonify({"error": str(e)}), 502


@app.route("/api/ai/parse-response", methods=["POST"])
def ai_parse_response():
    """Parse and validate JSON text returned by any AI (auto or manual paste)."""
    body = request.get_json(force=True)
    ptype = body.get("type", "")
    json_text = body.get("json_text", "").strip()

    if not json_text:
        return jsonify({"error": "json_text required"}), 400
    if ptype not in ("setup", "count", "invoice"):
        return jsonify({"error": "type must be setup, count, or invoice"}), 400

    # Strip markdown code fences if present
    json_text = re.sub(r"^```(?:json)?\s*", "", json_text, flags=re.MULTILINE)
    json_text = re.sub(r"\s*```$", "", json_text, flags=re.MULTILINE)
    json_text = json_text.strip()

    try:
        parsed = json.loads(json_text)
    except json.JSONDecodeError as e:
        return jsonify({"error": f"Invalid JSON: {e}", "raw": json_text[:500]}), 400

    bar = _load_bar()
    bottle_map = {b["id"]: b for b in _all_bottles(bar)}

    if ptype == "setup":
        stations = parsed.get("stations", [])
        if not isinstance(stations, list):
            return jsonify({"error": "Expected {stations: [...]}"}), 400
        # Assign IDs
        for s in stations:
            if not s.get("id"):
                s["id"] = _uid("stn-")
            for b in s.get("bottles", []):
                if not b.get("id"):
                    b["id"] = _uid("bot-")
        return jsonify({"stations": stations, "station_count": len(stations),
                        "bottle_count": sum(len(s.get("bottles", [])) for s in stations)})

    elif ptype == "count":
        matched = parsed.get("matched", [])
        unmatched = parsed.get("unmatched", [])
        # Validate bottle IDs exist
        valid = []
        for m in matched:
            bid = m.get("bottle_id", "")
            if bid in bottle_map:
                valid.append(m)
            else:
                # Try to fuzzy-match by name
                name = m.get("bottle_name", "").lower()
                found = next((b for b in bottle_map.values()
                              if b["name"].lower() == name), None)
                if found:
                    m["bottle_id"] = found["id"]
                    valid.append(m)
                else:
                    unmatched.append({"text": m.get("bottle_name", ""), "reason": "bottle not found in bar"})
        return jsonify({"matched": valid, "unmatched": unmatched})

    else:  # invoice
        items = parsed.get("items", [])
        if not isinstance(items, list):
            return jsonify({"error": "Expected {items: [...]}"}), 400
        return jsonify({"items": items, "item_count": len(items)})


# ─────────────────────────────────────────────
# ROUTES — Invoices
# ─────────────────────────────────────────────
@app.route("/api/invoice", methods=["POST"])
def save_invoice():
    """Save a reviewed invoice. Body: {date, items: [{bottle_name, size, qty, unit_cost, total}]}"""
    body = request.get_json(force=True)
    invoices = _load_invoices()
    record = {
        "id": _uid("inv-"),
        "date": body.get("date") or _now(),
        "items": body.get("items", []),
        "saved": _now(),
    }
    invoices.append(record)
    _save_invoices(invoices)
    print(f"[invoice] saved: {len(record['items'])} items")
    return jsonify(record), 201


@app.route("/api/invoices", methods=["GET"])
def get_invoices():
    invoices = _load_invoices()
    return jsonify(sorted(invoices, key=lambda x: x.get("date", ""), reverse=True))


# ─────────────────────────────────────────────
# ROUTES — Financial
# ─────────────────────────────────────────────
@app.route("/api/usage", methods=["GET"])
def get_usage():
    """Per-bottle usage: previous_level + invoiced_qty_since_prev_count - current_level."""
    counts = _load_counts()
    if len(counts) < 2:
        return jsonify({"error": "need at least 2 counts for usage", "usage": []}), 200

    current = counts[-1]
    previous = counts[-2]
    prev_date = previous.get("date", "")

    cur_levels = {e["bottle_id"]: float(e.get("level", 0)) for e in current.get("entries", [])}
    prev_levels = {e["bottle_id"]: float(e.get("level", 0)) for e in previous.get("entries", [])}

    # Sum invoice quantities received since previous count
    invoices = _load_invoices()
    inv_since = [inv for inv in invoices if inv.get("date", "") > prev_date]
    inv_qty_map = {}
    for inv in inv_since:
        for item in inv.get("items", []):
            name = item.get("bottle_name", "").lower()
            qty = float(item.get("qty") or 0)
            inv_qty_map[name] = inv_qty_map.get(name, 0) + qty

    bar = _load_bar()
    usage = []
    for bottle in _all_bottles(bar):
        bid = bottle["id"]
        if bid not in cur_levels or bid not in prev_levels:
            continue
        cur = cur_levels[bid]
        prev = prev_levels[bid]
        inv_qty = inv_qty_map.get(bottle["name"].lower(), 0)
        used = round(prev + inv_qty - cur, 3)
        cost = bottle.get("cost", 0)
        usage.append({
            "bottle_id": bid,
            "bottle_name": bottle["name"],
            "previous_level": prev,
            "invoiced_qty": inv_qty,
            "current_level": cur,
            "usage": used,
            "cost_impact": round(used * cost, 2),
        })

    usage.sort(key=lambda x: x["usage"], reverse=True)
    return jsonify({"usage": usage, "period_start": prev_date, "period_end": current.get("date", "")})


def _parse_size_ml(size_str):
    """Parse bottle size string to ml. Returns float or None."""
    if not size_str:
        return None
    s = str(size_str).strip().lower().replace(",", "")
    # "1.75l", "1l", "750ml", "375ml", "200ml"
    m = re.match(r"(\d+(?:\.\d+)?)\s*(l|ml|liter|litre)?", s)
    if not m:
        return None
    val = float(m.group(1))
    unit = (m.group(2) or "ml").lower()
    if unit in ("l", "liter", "litre"):
        return val * 1000
    return val  # ml


@app.route("/api/cost-per-oz", methods=["GET"])
def get_cost_per_oz():
    bar = _load_bar()
    result = []
    for bottle in _all_bottles(bar):
        cost = bottle.get("cost", 0)
        size_ml = _parse_size_ml(bottle.get("size", "750ml"))
        if size_ml and cost:
            oz = size_ml / 29.5735
            cpo = round(cost / oz, 4)
        else:
            cpo = None
        result.append({
            "bottle_id": bottle["id"],
            "bottle_name": bottle["name"],
            "category": bottle.get("category", ""),
            "size": bottle.get("size", ""),
            "size_ml": size_ml,
            "unit_cost": cost,
            "cost_per_oz": cpo,
        })
    result.sort(key=lambda x: (x["category"], x["bottle_name"]))
    return jsonify(result)


@app.route("/api/export", methods=["GET"])
def export_xlsx():
    """Generate and return an XLSX workbook of current inventory + recent counts."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "openpyxl not installed. Run: pip3 install openpyxl"}), 500

    bar = _load_bar()
    counts = _load_counts()
    bottles = _all_bottles(bar)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Inventory ──
    ws1 = wb.active
    ws1.title = "Inventory"
    hdr_font = Font(bold=True, color="FFE0E6EF")
    hdr_fill = PatternFill("solid", fgColor="FF16202E")
    thin = Side(border_style="thin", color="FF2A3A4A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Product", "Category", "Station", "Size", "Par", "Level", "Unit Cost", "Total Value"]
    widths = [30, 14, 20, 10, 8, 8, 12, 14]
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws1.cell(row=1, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Bottles are nested under stations; walk them in order so each carries its station name
    r = 2
    for station in bar.get("stations", []):
        for b in station.get("bottles", []):
            vals = [
                b.get("name", ""),
                b.get("category", ""),
                station.get("name", ""),
                b.get("size", ""),
                b.get("par_level", b.get("par", 1)),
                b.get("current_level", b.get("level", 0)),
                b.get("cost", 0),
            ]
            for ci, v in enumerate(vals, 1):
                cell = ws1.cell(row=r, column=ci, value=v)
                cell.border = border
            # Total Value formula
            tv = ws1.cell(row=r, column=8, value=f"=G{r}*F{r}")
            tv.border = border
            tv.number_format = '"$"#,##0.00'
            r += 1

    ws1.freeze_panes = "A2"

    # ── Sheet 2: Count History ──
    ws2 = wb.create_sheet("Count History")
    for i, h in enumerate(["Date", "Product", "Level", "Previous Level", "Change"], 1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center")
        c.border = border

    bottle_name_map = {b["id"]: b["name"] for b in bottles}
    row = 2
    for cnt in reversed(counts[-10:]):
        date_str = cnt.get("date", "")[:10]
        for entry in cnt.get("entries", []):
            bid = entry.get("bottle_id", "")
            ws2.cell(row=row, column=1, value=date_str).border = border
            ws2.cell(row=row, column=2, value=bottle_name_map.get(bid, bid)).border = border
            ws2.cell(row=row, column=3, value=entry.get("level", "")).border = border
            ws2.cell(row=row, column=4, value="").border = border
            ws2.cell(row=row, column=5, value=f"=C{row}-D{row}").border = border
            row += 1

    ws2.freeze_panes = "A2"

    # Write to buffer and return
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    bar_name = bar.get("bar_name", "Bar").replace(" ", "-") or "Bar"
    filename = f"{bar_name}-Inventory-{datetime.now().strftime('%Y%m%d')}.xlsx"

    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


# ─────────────────────────────────────────────
# ROUTES — Stats & Variance
# ─────────────────────────────────────────────
@app.route("/api/stats", methods=["GET"])
def get_stats():
    bar = _load_bar()
    bottles = _all_bottles(bar)
    counts = _load_counts()

    total_products = len(bottles)
    total_value = sum(b.get("cost", 0) * b.get("current_level", 0) for b in bottles)
    below_par = sum(1 for b in bottles if b.get("current_level", 0) < b.get("par_level", 1))
    last_count = counts[-1]["date"] if counts else None

    return jsonify({
        "total_products": total_products,
        "total_value": round(total_value, 2),
        "below_par": below_par,
        "last_count_date": last_count,
        "total_counts": len(counts),
        "station_count": len(bar.get("stations", [])),
        "bar_name": bar.get("bar_name", ""),
    })


@app.route("/api/variance", methods=["GET"])
def get_variance():
    """Compare last two counts and return variance per bottle."""
    counts = _load_counts()
    if len(counts) < 2:
        return jsonify({"error": "need at least 2 counts for variance", "variance": []}), 200

    current = counts[-1]
    previous = counts[-2]

    # Build lookup: bottle_id -> level for each count
    cur_levels = {e["bottle_id"]: e["level"] for e in current.get("entries", [])}
    prev_levels = {e["bottle_id"]: e["level"] for e in previous.get("entries", [])}

    bar = _load_bar()
    variance = []
    for bottle in _all_bottles(bar):
        bid = bottle["id"]
        if bid in cur_levels and bid in prev_levels:
            cur = float(cur_levels[bid])
            prev = float(prev_levels[bid])
            diff = cur - prev
            variance.append({
                "bottle_id": bid,
                "bottle_name": bottle["name"],
                "previous_level": prev,
                "current_level": cur,
                "variance": round(diff, 3),
                "cost_impact": round(diff * bottle.get("cost", 0), 2),
            })

    # Sort by variance (biggest drops first)
    variance.sort(key=lambda v: v["variance"])

    return jsonify({
        "current_count": current["id"],
        "current_date": current["date"],
        "previous_count": previous["id"],
        "previous_date": previous["date"],
        "variance": variance,
    })


# ─────────────────────────────────────────────
# ROUTES — Utility
# ─────────────────────────────────────────────
@app.route("/ping")
def ping():
    return jsonify({"status": "ok", "app": "bar-inventory", "port": 5051})


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
if __name__ == "__main__":
    # Init data files if missing
    if not os.path.exists(_BAR_FILE):
        _save_bar({"bar_name": "", "created": "", "stations": []})
        print(f"[init] created {_BAR_FILE}")
    if not os.path.exists(_COUNT_FILE):
        _save_counts([])
        print(f"[init] created {_COUNT_FILE}")
    if not os.path.exists(_CONFIG_FILE):
        _save_config({"api_base_url": "", "api_key": "", "model": ""})
        print(f"[init] created {_CONFIG_FILE}")
    if not os.path.exists(_INVOICE_FILE):
        _save_invoices([])
        print(f"[init] created {_INVOICE_FILE}")

    print("\n  Bar Inventory — Open Source Barware")
    print("  ──────────────────────────────────────────────────────")
    print(f"  Dashboard:  http://localhost:5051/")
    print(f"  API:        http://localhost:5051/api/bar")
    print(f"  Stats:      http://localhost:5051/api/stats")
    print(f"  Ping:       http://localhost:5051/ping")
    print(f"  ──────────────────────────────────────────────────────")
    print(f"  Data:       {_BAR_FILE}")
    print(f"  Counts:     {_COUNT_FILE}")

    # SSL — same pattern as OVLP pop
    cert = os.path.join(_DIR, "localhost+1.pem")
    key = os.path.join(_DIR, "localhost+1-key.pem")
    ssl_ctx = (cert, key) if os.path.exists(cert) else None
    if ssl_ctx:
        print("  HTTPS: using mkcert cert (trusted by Chrome)\n")
    else:
        print("  Running HTTP (no cert found)\n")

    app.run(host="0.0.0.0", port=5051, debug=False, threaded=True, ssl_context=ssl_ctx)
